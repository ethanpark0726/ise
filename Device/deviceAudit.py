import sys
import os
import subprocess
import openpyxl
import datetime

from ise import ERS  # noqa E402
from pprint import pprint  # noqa E402
from config import uri, endpoint, endpoint_group, user, identity_group, device, device_group, trustsec  # noqa E402
from exchangelib import *
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

def getList(iseObj):

    pageNumber = 1
    deviceList = list()

    temp = iseObj.get_devices(page=pageNumber)
    deviceList.append(temp.get('response'))

    while temp.get('nextPage'):
        pageNumber += 1
        temp = iseObj.get_devices(page=pageNumber)
        deviceList.append(temp.get('response'))
    
    print("--- List gathering Complete!")
    return deviceList

def getDeviceIDList(deviceList):

    deviceIDList = list()
    for i in range(len(deviceList)):
        for j in range(len(deviceList[i])):
            deviceIDList.append(deviceList[i][j][0])
    print("--- Device ID gathering Complete!")
    return deviceIDList

def getIPList(iseObj, deviceIDList):
    
    ipList = list()

    for deviceID in deviceIDList:
        ipList.append(iseObj.get_device(deviceID).get('response').get('NetworkDeviceIPList')[0].get('ipaddress'))
    print("--- Device IP gathering Complete!")
    return ipList

def getPingResult(deviceIPList):

    pingList = dict()

    for deviceIP in deviceIPList:
        
        result = os.system('ping -n 1 -w 500 ' + deviceIP)

        if result == 0:
            pingList[deviceIP] = 'Okay'
        else:
            pingList[deviceIP] = 'Needs to check'
    print("--- Device Ping list gathering Complete!")
    return pingList

def createExcelFile():
    # Excel File Creation
    nowDate = 'Report Date: ' + str(datetime.datetime.now().strftime('%Y-%m-%d'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ISE Device Audit'
    
    # Pretty display for the File
    font = Font(bold=True)
    alignment = Alignment(horizontal='center')
    bgColor = PatternFill(fgColor='BFBFBFBF', patternType='solid')
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    ws['A2'] = nowDate
    ws['A4'] = 'Hostname'
    ws['A4'].alignment = alignment
    ws['A4'].font = font
    ws['A4'].fill = bgColor
    ws['A4'].border = border

    ws['B4'] = 'IP Address'
    ws['B4'].alignment = alignment
    ws['B4'].font = font
    ws['B4'].fill = bgColor
    ws['B4'].border = border

    ws['C4'] = 'Ping Status'
    ws['C4'].alignment = alignment
    ws['C4'].font = font  
    ws['C4'].fill = bgColor
    ws['C4'].border = border

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    fileName = 'ISE_Device_Audit.xlsx'
    wb.save(fileName)
    wb.close()

def saveExcelFile(deviceIDList, deviceIPList, pingResult):

    fileName = 'ISE_Device_Audit.xlsx'
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

    cellNumber = 5

    for i in range(len(deviceIDList)):
        ws['A' + str(cellNumber)] = deviceIDList[i]
        ws['A' + str(cellNumber)].alignment = alignment
        ws['A' + str(cellNumber)].border = border

        ws['B' + str(cellNumber)] = deviceIPList[i]
        ws['B' + str(cellNumber)].alignment = alignment
        ws['B' + str(cellNumber)].border = border

        ws['C' + str(cellNumber)] = pingResult.get(deviceIPList[i])
        ws['C' + str(cellNumber)].alignment = alignment
        ws['C' + str(cellNumber)].border = border

        cellNumber += 1

    wb.save('ISE_Device_Audit.xlsx')
    wb.close()

if __name__ == "__main__":

    ise = ERS(ise_node=uri['ise_node'], ers_user=uri['ers_user'], ers_pass=uri['ers_pass'], verify=False,
          disable_warnings=True, timeout=15)
    
    createExcelFile()

    deviceList = getList(ise)
    
    deviceIDList = getDeviceIDList(deviceList)
    
    deviceIPList = getIPList(ise, deviceIDList)
    pingResult = getPingResult(deviceIPList)

    saveExcelFile(deviceIDList, deviceIPList, pingResult)

    credentials = Credentials(username='hpar0001', password='Rhakdnj25~!')

    account = Account(primary_smtp_address='hpar0001@shands.ufl.edu', credentials=credentials, autodiscover=True, access_type=DELEGATE)

    bodyContents = """ISE's device list audit has been conducted.
    Please check the attachment"""
    m = Message(account=account, subject='[ISE-Audit] ISE Device Audit', body=bodyContents, 
    to_recipients=[
        Mailbox(email_address='ahc_neteng@shands.ufl.edu')])
    
    filePath = open('P:/Script/ISE Audit/ISE_Device_Audit.xlsx', 'rb').read()
    m.attach(FileAttachment(name='ISE_Device_Audit.xlsx', content=filePath))
    m.send()
